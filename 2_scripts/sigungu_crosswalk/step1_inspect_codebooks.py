"""Phase 1-A Step 1: 1997-1999 시군구 코드집 구조 점검.

산출 없음. stdout에 시트명·헤더·첫 20행 출력만.
표준화 스크립트는 이 결과를 보고 별도 작성.
"""
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(r"C:/Users/82103/Downloads/trade_mortality_korea")
RAW_DIR = ROOT / "0_raw" / "mortality_kostat" / "usrcnfrm"

FILES = {
    1997: "시군구코드집(공공용)_사망원인통계_사망_연간자료_B형(제공)_1997.xlsx",
    1998: "시군구코드집(공공용)_사망원인통계_사망_연간자료_B형(제공)_1998.xlsx",
    1999: "시군구코드집(공공용)_사망원인통계_사망_연간자료_B형(제공)_1999.xlsx",
}


def inspect(year: int, fname: str) -> None:
    path = RAW_DIR / fname
    print(f"\n{'='*80}\n{year} :: {fname}\n{'='*80}")
    if not path.exists():
        print(f"  ❌ 파일 없음: {path}")
        return
    print(f"  파일 크기: {path.stat().st_size:,} bytes")

    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"  시트 수: {len(wb.sheetnames)}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"    - 시트 '{sn}' :: dim={ws.calculate_dimension()}, max_row={ws.max_row}, max_col={ws.max_column}")
    wb.close()

    for sn in wb.sheetnames:
        try:
            df_head = pd.read_excel(path, sheet_name=sn, header=None, nrows=20, dtype=str)
        except Exception as e:
            print(f"\n  ⚠ 시트 '{sn}' 읽기 실패: {e}")
            continue
        print(f"\n  --- 시트 '{sn}' 첫 20행 (header=None) ---")
        print(df_head.to_string(max_cols=12, max_colwidth=30))


def main() -> None:
    for y, f in FILES.items():
        inspect(y, f)
    print("\n" + "="*80 + "\n점검 완료. 위 출력 보고 표준화 로직 결정.")


if __name__ == "__main__":
    main()
