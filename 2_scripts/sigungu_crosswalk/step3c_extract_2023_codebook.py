"""
Step 3c — KOSTAT 2023 사망원인통계 파일설계서 → 시군구 코드집 추출

Input:
 0_raw/mortality_kostat/usrcnfrm/파일설계서(공공용)_사망원인통계_사망연간자료B형(제공)_2023(코드집포함).xlsx
 Sheet: '코드정보' (742 rows, 5 cols)
 Columns: 코드번호 | 항목명 | 코드 | 코드의미 및 설명 | 특이사항

Filter rule:
 forward-fill 항목명 (merged cells), keep only rows where 항목명 == '사망자주소행정구역시군구코드'
 → 261 rows expected

Output:
 3_derived/sigungu/step3c_codebook_2023.csv
 columns: year, sido_code, sido_name, raw_code, sigungu_name, code_len
 schema matches step1_codebook_old.csv
"""
from __future__ import annotations
import sys
import io
from pathlib import Path

import openpyxl
import pandas as pd

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve.parents[2]
SRC = ROOT / "0_raw" / "mortality_kostat" / "usrcnfrm" / (
 "파일설계서(공공용)_사망원인통계_사망연간자료B형(제공)_2023(코드집포함).xlsx"
)
OUT = ROOT / "3_derived" / "sigungu" / "step3c_codebook_2023.csv"

ITEM_TARGET = "사망자주소행정구역시군구코드"

def main -> None:
 if not SRC.exists:
 raise FileNotFoundError(f"missing source xlsx: {SRC}")

 wb = openpyxl.load_workbook(SRC, data_only=True)
 if "코드정보" not in wb.sheetnames:
 raise KeyError(f"sheet '코드정보' not found in {SRC.name} (sheets={wb.sheetnames})")
 ws = wb["코드정보"]

 rows: list[tuple[str, str]] = 
 current_item: str | None = None
 for i, row in enumerate(ws.iter_rows(values_only=True)):
 if i < 2:
 continue
 _code_no, item_name, code, meaning, _note = row
 if item_name:
 current_item = str(item_name).strip
 if current_item!= ITEM_TARGET:
 continue
 if code is None:
 continue
 code_s = str(code).strip
 meaning_s = (str(meaning).strip if meaning is not None else "")
 rows.append((code_s, meaning_s))

 if not rows:
 raise RuntimeError("no rows extracted — check sheet layout / item name")

 df = pd.DataFrame(rows, columns=["raw_code", "full_name"])
 if df["raw_code"].duplicated.any:
 dups = df[df["raw_code"].duplicated(keep=False)]
 raise RuntimeError(f"duplicate raw_code in 2023 codebook:\n{dups}")

 df["sido_code"] = df["raw_code"].str[:2].astype(int)
 df["code_len"] = df["raw_code"].str.len

 parts = df["full_name"].str.split(" ", n=1, expand=True)
 parts.columns = ["sido_name", "sigungu_name"]
 df["sido_name"] = parts["sido_name"]
 df["sigungu_name"] = parts["sigungu_name"]
 df["year"] = 2023

 out = df[["year", "sido_code", "sido_name", "raw_code", "sigungu_name", "code_len"]].copy
 out = out.sort_values(["sido_code", "raw_code"]).reset_index(drop=True)

 OUT.parent.mkdir(parents=True, exist_ok=True)
 out.to_csv(OUT, index=False, encoding="utf-8-sig")

 print(f"[step3c] rows extracted: {len(out)}")
 print(f"[step3c] distinct sido: {sorted(out['sido_code'].unique.tolist)}")
 print(f"[step3c] code_len min/max: {out['code_len'].min}/{out['code_len'].max}")
 print(f"[step3c] sido name set:")
 print(out.drop_duplicates("sido_code")[["sido_code", "sido_name"]].to_string(index=False))
 print(f"[step3c] wrote: {OUT}")

if __name__ == "__main__":
 main
