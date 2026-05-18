"""
KSIC 8차 → 9차 → 10차 chain merge.

보유 매핑표:
 8차_9차개정 연계표.xls (col 0 = KSIC 8, col 2 = KSIC 9)
 KSIC연계표(9차_10차).xlsx (col 0 = KSIC 10, col 2 = KSIC 9)
 한국표준산업분류 제11차-제10차 연계표.xlsx (sheet 신구연계표 col 0 = KSIC 11, col 2 = KSIC 10)

산출:
 3_derived/sigungu/ksic_chain_lookup.csv ← researchall (KSIC 10) 와 매칭 가능한 단일 lookup
 columns: ksic_8, ksic_9, ksic_10, ksic_11
 + 연도별 차수 매핑 (1994-2006: 8차 / 2007-2016: 9차 / 2017-2023: 10차 / 2024+: 11차)
"""
from __future__ import annotations
import sys, csv
from pathlib import Path
import openpyxl, xlrd

sys.path.insert(0, str(Path(__file__).resolve.parents[1]))
from lib.config import RAW_DIR, DERIVED_DIR

if hasattr(sys.stdout, "reconfigure"):
 sys.stdout.reconfigure(encoding="utf-8", errors="replace")

CHN = RAW_DIR / "kosis_business_survey/ksic_version_concordance"
OUT = DERIVED_DIR / "sigungu/ksic_chain_lookup.csv"

def extract(values):
 s = set
 for v in values:
 v = str(v or '').strip.split('.')[0]
 if v.isdigit and len(v) == 5:
 s.add(v)
 return s

def main:
 OUT.parent.mkdir(parents=True, exist_ok=True)

 # 8↔9 mapping (xls)
 wb = xlrd.open_workbook(str(CHN / "8차_9차개정 연계표.xls"))
 sh = wb.sheet_by_index(0)
 map_8_to_9 = {} # 8 → set of 9
 for r in range(sh.nrows):
 c8 = str(sh.cell_value(r, 0)).strip.split('.')[0]
 c9 = str(sh.cell_value(r, 2)).strip.split('.')[0]
 if c8.isdigit and len(c8) == 5 and c9.isdigit and len(c9) == 5:
 map_8_to_9.setdefault(c8, set).add(c9)

 # 9↔10 mapping (xlsx)
 wb = openpyxl.load_workbook(str(CHN / "KSIC연계표(9차_10차).xlsx"), read_only=True, data_only=True)
 rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
 map_9_to_10 = {}
 for r in rows[2:]:
 c10 = str(r[0] or '').strip.split('.')[0]
 c9 = str(r[2] or '').strip.split('.')[0]
 if c9.isdigit and len(c9) == 5 and c10.isdigit and len(c10) == 5:
 map_9_to_10.setdefault(c9, set).add(c10)

 # 10↔11 mapping (xlsx 신구연계표)
 wb = openpyxl.load_workbook(str(CHN / "한국표준산업분류 제11차-제10차 연계표.xlsx"), read_only=True, data_only=True)
 rows = list(wb['신구연계표'].iter_rows(values_only=True))
 map_10_to_11 = {}
 for r in rows[2:]:
 c11 = str(r[0] or '').strip.split('.')[0]
 c10 = str(r[2] or '').strip.split('.')[0]
 if c10.isdigit and len(c10) == 5 and c11.isdigit and len(c11) == 5:
 map_10_to_11.setdefault(c10, set).add(c11)

 # Chain join: 8 → 9 → 10 → 11
 chain_rows = 
 for c8, c9_set in map_8_to_9.items:
 for c9 in c9_set:
 c10_set = map_9_to_10.get(c9, {""})
 for c10 in c10_set:
 c11_set = map_10_to_11.get(c10, {""})
 for c11 in c11_set:
 chain_rows.append({"ksic_8": c8, "ksic_9": c9, "ksic_10": c10, "ksic_11": c11})

 # Rows for KSIC 9 codes that don't appear as upstream of 8 (new in 9차)
 seen_9 = {r["ksic_9"] for r in chain_rows}
 for c9, c10_set in map_9_to_10.items:
 if c9 in seen_9: continue
 for c10 in c10_set:
 for c11 in map_10_to_11.get(c10, {""}):
 chain_rows.append({"ksic_8": "", "ksic_9": c9, "ksic_10": c10, "ksic_11": c11})

 # KSIC 10 unique (new in 10차)
 seen_10 = {r["ksic_10"] for r in chain_rows}
 for c10, c11_set in map_10_to_11.items:
 if c10 in seen_10: continue
 for c11 in c11_set:
 chain_rows.append({"ksic_8": "", "ksic_9": "", "ksic_10": c10, "ksic_11": c11})

 with OUT.open("w", encoding="utf-8-sig", newline="") as f:
 writer = csv.DictWriter(f, fieldnames=["ksic_8", "ksic_9", "ksic_10", "ksic_11"])
 writer.writeheader
 writer.writerows(chain_rows)

 # Summary
 print(f"[ksic chain] {len(chain_rows):,} rows -> {OUT}")
 distinct = {k: len({r[k] for r in chain_rows if r[k]}) for k in ["ksic_8", "ksic_9", "ksic_10", "ksic_11"]}
 print(f" distinct: {distinct}")
 print(f" 연도-차수 매핑 권장:")
 print(f" 1994-2006: KSIC 8차")
 print(f" 2007-2016: KSIC 9차")
 print(f" 2017-2023: KSIC 10차 (researchall HS-KSIC 와 직접 일치)")
 print(f" 2024+: KSIC 11차")

if __name__ == "__main__":
 main
